"""Экспорт в Excel"""

from openpyxl import Workbook


class ExportService:
    def export_to_excel(self, requirements, links, file_path):
        """Экспорт списка требований и связей в один .xlsx."""
        wb = Workbook()

        # --- Лист 1: требования ---
        ws_req = wb.active
        ws_req.title = "Requirements"
        ws_req.append([
            "id",
            "title",
            "type",
            "status",
            "priority",
            "source",
            "author",
        ])

        for req in requirements:
            d = req.to_dict()
            ws_req.append([
                d.get("id"),
                d.get("title"),
                d.get("requirement_type"),
                d.get("status"),
                d.get("priority"),
                d.get("source"),
                d.get("author"),
            ])

        # --- Лист 2: связи ---
        ws_links = wb.create_sheet("Links")
        ws_links.append(["id", "source_id", "target_id", "link_type"])

        for link in links:
            ws_links.append([
                getattr(link, "id", None),
                link.source_requirement_id,
                link.target_requirement_id,
                link.link_type.value,
            ])

        wb.save(file_path)

    def export_matrix_to_excel(self, requirements, links, file_path):
        """Экспорт матрицы пересечений (кто кого покрывает/зависит/...)"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Matrix"

        req_ids = [r.id for r in requirements]

        # Собираем быстрый словарь: matrix[source][target] = "тип"
        matrix = {}
        for l in links:
            matrix.setdefault(l.source_requirement_id, {})[l.target_requirement_id] = l.link_type.value

        # Шапка
        ws.append([""] + [f"#{rid}" for rid in req_ids])

        # Строки
        for source_id in req_ids:
            row = [f"#{source_id}"]
            for target_id in req_ids:
                if source_id == target_id:
                    row.append("-")
                else:
                    row.append(matrix.get(source_id, {}).get(target_id, ""))
            ws.append(row)

        wb.save(file_path)
